"""
excel_exporter.py — GT Scout Bot
Exportações separadas por aba/contexto.
"""

import os
import io
import smtplib
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
BR_TZ = pytz.timezone("America/Sao_Paulo")


def now_br():
    return datetime.now(BR_TZ)


# ── Estilos ────────────────────────────────────────────────────
FILL_WIN    = PatternFill("solid", fgColor="1A7A4A")
FILL_LOSE   = PatternFill("solid", fgColor="8B1A1A")
FILL_DRAW   = PatternFill("solid", fgColor="7A6A00")
FILL_HEADER = PatternFill("solid", fgColor="0D1117")
FILL_ROW1   = PatternFill("solid", fgColor="161B22")
FILL_ROW2   = PatternFill("solid", fgColor="0D1117")

FONT_WIN    = Font(bold=True, color="00FF88", name="Calibri", size=11)
FONT_LOSE   = Font(bold=True, color="FF6B6B", name="Calibri", size=11)
FONT_DRAW   = Font(bold=True, color="FFD700", name="Calibri", size=11)
FONT_HEADER = Font(bold=True, color="00E5A0", name="Calibri", size=11)
FONT_NORMAL = Font(color="C9D1D9", name="Calibri", size=10)
FONT_MUTED  = Font(color="8B949E", name="Calibri", size=10)
FONT_BOLD   = Font(bold=True, color="E6EDF3", name="Calibri", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")


def _border():
    s = Side(style="thin", color="21262D")
    return Border(left=s, right=s, top=s, bottom=s)


def _frow(ri):
    return FILL_ROW1 if ri % 2 == 0 else FILL_ROW2


def _header_row(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _cell(ws, row, col, value, fill, font, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = font
    c.border = _border()
    c.alignment = align or ALIGN_CENTER
    return c


# ──────────────────────────────────────────────────────────────
# 1. Relatório de partidas
# ──────────────────────────────────────────────────────────────
def build_excel_reports(matches) -> bytes:
    """Planilha com partidas coletadas (finalizadas)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Partidas Coletadas"

    headers = [
        "#", "Data", "Semana", "Temporada",
        "Player Casa", "Gols Casa",
        "Gols Visitante", "Player Visitante",
        "Resultado"
    ]
    widths = [5, 18, 7, 28, 18, 11, 11, 18, 22]
    _header_row(ws, headers, widths)

    for ri, m in enumerate(matches, 2):
        hs, as_ = m.home_score, m.away_score
        fill = _frow(ri)

        if hs is None or as_ is None:
            result_str = "—"
            fh = fa = fill; fnh = fna = FONT_MUTED
        elif hs > as_:
            result_str = f"{m.home_nickname} venceu"
            fh, fnh = FILL_WIN, FONT_WIN
            fa, fna = FILL_LOSE, FONT_LOSE
        elif hs < as_:
            result_str = f"{m.away_nickname} venceu"
            fh, fnh = FILL_LOSE, FONT_LOSE
            fa, fna = FILL_WIN, FONT_WIN
        else:
            result_str = "Empate"
            fh = fa = FILL_DRAW; fnh = fna = FONT_DRAW

        kickoff = (m.kickoff or "")[:16].replace("T", " ")

        row = [
            ri - 1, kickoff, m.week or "", m.season_name or "",
            m.home_nickname or "", hs if hs is not None else "",
            as_ if as_ is not None else "", m.away_nickname or "",
            result_str
        ]

        for ci, val in enumerate(row, 1):
            if ci == 5:
                _cell(ws, ri, ci, val, fh, fnh)
            elif ci == 8:
                _cell(ws, ri, ci, val, fa, fna)
            elif ci == 9:
                if "venceu" in result_str:
                    _cell(ws, ri, ci, val, FILL_WIN, FONT_WIN)
                elif result_str == "Empate":
                    _cell(ws, ri, ci, val, FILL_DRAW, FONT_DRAW)
                else:
                    _cell(ws, ri, ci, val, fill, FONT_MUTED)
            else:
                _cell(ws, ri, ci, val, fill, FONT_MUTED if ci in (1, 2, 3) else FONT_NORMAL)
        ws.row_dimensions[ri].height = 20

    return _to_bytes(wb)


# ──────────────────────────────────────────────────────────────
# 2. Estatísticas individuais
# ──────────────────────────────────────────────────────────────
def build_excel_stats(stats_list) -> bytes:
    """Planilha com estatísticas individuais dos players."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estatísticas"

    headers = [
        "#", "Player", "PJ", "V", "E", "D",
        "GF", "GC", "Saldo",
        "Méd. GF/P", "Méd. GC/P", "Méd. Total/P"
    ]
    widths = [5, 20, 6, 6, 6, 6, 7, 7, 8, 12, 12, 14]
    _header_row(ws, headers, widths)

    for ri, s in enumerate(stats_list, 2):
        gp = s.get("games_played", 0)
        gf = s.get("goals_for", 0)
        ga = s.get("goals_against", 0)
        w  = s.get("wins", 0)
        d  = s.get("draws", 0)
        l  = s.get("losses", 0)
        avg_f = round(gf / gp, 2) if gp else 0
        avg_a = round(ga / gp, 2) if gp else 0
        avg_t = round((gf + ga) / gp, 2) if gp else 0
        saldo = gf - ga
        fill = _frow(ri)

        row = [ri - 1, s.get("nickname", ""), gp, w, d, l,
               gf, ga, saldo, avg_f, avg_a, avg_t]

        for ci, val in enumerate(row, 1):
            if ci == 2:
                _cell(ws, ri, ci, val, fill, FONT_BOLD)
            elif ci == 10:
                _cell(ws, ri, ci, val, fill,
                      Font(bold=True, color="00FF88", name="Calibri", size=10))
            elif ci == 11:
                _cell(ws, ri, ci, val, fill,
                      Font(bold=True, color="FF6B6B", name="Calibri", size=10))
            elif ci == 12:
                _cell(ws, ri, ci, val, fill,
                      Font(bold=True, color="8B949E", name="Calibri", size=10))
            else:
                _cell(ws, ri, ci, val, fill, FONT_MUTED if ci == 1 else FONT_NORMAL)
        ws.row_dimensions[ri].height = 20

    return _to_bytes(wb)


# ──────────────────────────────────────────────────────────────
# 3. Head-to-Head
# ──────────────────────────────────────────────────────────────
def build_excel_h2h(games, p1, p2, p1_stats, p2_stats) -> bytes:
    """Planilha com confrontos diretos entre dois players."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{p1} vs {p2}"

    # Resumo no topo
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value=f"Confronto Direto: {p1} vs {p2}")
    title_cell.fill = FILL_HEADER
    title_cell.font = Font(bold=True, color="00E5A0", name="Calibri", size=13)
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    # Linha de resumo dos players
    summary_headers = ["", "PJ", "V", "E", "D", "GF", "GC", "Saldo", "Méd. GF/P"]
    summary_widths   = [20, 6, 6, 6, 6, 6, 6, 8, 12]
    for ci, (h, w) in enumerate(zip(summary_headers, summary_widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER; c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (pname, ps) in enumerate([(p1, p1_stats), (p2, p2_stats)], 3):
        fill = FILL_WIN if pname == p1 else FILL_LOSE
        vals = [
            pname,
            ps["n"], ps["wins"], ps["draws"], ps["losses"],
            ps["gf"], ps["ga"], ps["gf"] - ps["ga"],
            round(ps["gf"] / ps["n"], 2) if ps["n"] else 0
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill if ci == 1 else _frow(ri)
            c.font = FONT_WIN if ci == 1 else FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = _border()
        ws.row_dimensions[ri].height = 20

    # Espaço
    ws.row_dimensions[5].height = 8

    # Cabeçalho da lista de confrontos
    match_headers = ["#", "Data", "Temporada",
                     p1, "Gols Casa", "Gols Visit.", p2, "Vencedor"]
    match_widths   = [5, 18, 28, 18, 11, 11, 18, 22]
    for ci, (h, w) in enumerate(zip(match_headers, match_widths), 1):
        c = ws.cell(row=6, column=ci, value=h)
        c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER; c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[6].height = 22
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False

    for ri, g in enumerate(games, 7):
        fill = _frow(ri)
        hs, as_ = g["p1_score"], g["p2_score"]

        if g["winner"] == p1:
            fh, fnh = FILL_WIN, FONT_WIN
            fa, fna = FILL_LOSE, FONT_LOSE
            fw = FONT_WIN
        elif g["winner"] == p2:
            fh, fnh = FILL_LOSE, FONT_LOSE
            fa, fna = FILL_WIN, FONT_WIN
            fw = FONT_WIN
        else:
            fh = fa = FILL_DRAW; fnh = fna = fw = FONT_DRAW

        kickoff = (g.get("kickoff", "") or "")[:16].replace("T", " ")
        vals = [ri - 6, kickoff, g.get("season_name", ""),
                p1, hs, as_, p2, g["winner"]]

        for ci, val in enumerate(vals, 1):
            if ci == 4:
                _cell(ws, ri, ci, val, fh, fnh)
            elif ci == 7:
                _cell(ws, ri, ci, val, fa, fna)
            elif ci == 8:
                _cell(ws, ri, ci, val, FILL_WIN if g["winner"] not in ("Empate",) else FILL_DRAW, fw)
            else:
                _cell(ws, ri, ci, val, fill, FONT_MUTED if ci in (1, 2) else FONT_NORMAL)
        ws.row_dimensions[ri].height = 20

    return _to_bytes(wb)


# ──────────────────────────────────────────────────────────────
# 4. Dados para gráficos (mesmos dados de estatísticas)
# ──────────────────────────────────────────────────────────────
def build_excel_charts(stats_list) -> bytes:
    """Planilha com dados usados nos gráficos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados dos Gráficos"

    headers = [
        "Player", "PJ",
        "Méd. GF/P", "Méd. GC/P", "Méd. Total/P",
        "GF Total", "GC Total", "Saldo"
    ]
    widths = [20, 6, 14, 14, 16, 10, 10, 10]
    _header_row(ws, headers, widths)

    for ri, s in enumerate(stats_list, 2):
        gp = s.get("games_played", 0)
        gf = s.get("goals_for", 0)
        ga = s.get("goals_against", 0)
        avg_f = round(gf / gp, 2) if gp else 0
        avg_a = round(ga / gp, 2) if gp else 0
        avg_t = round((gf + ga) / gp, 2) if gp else 0
        fill = _frow(ri)

        row = [s.get("nickname", ""), gp, avg_f, avg_a, avg_t, gf, ga, gf - ga]
        for ci, val in enumerate(row, 1):
            _cell(ws, ri, ci, val, fill, FONT_BOLD if ci == 1 else FONT_NORMAL)
        ws.row_dimensions[ri].height = 20

    return _to_bytes(wb)


# ──────────────────────────────────────────────────────────────
# Compat: build_excel original → agora chama reports
# ──────────────────────────────────────────────────────────────
def build_excel(matches) -> bytes:
    return build_excel_reports(matches)


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────────────────────
# Email semanal
# ──────────────────────────────────────────────────────────────
def send_excel_email(matches, recipient: str = None):
    user      = os.getenv("EMAIL_USER", "")
    password  = os.getenv("EMAIL_PASSWORD", "")
    smtp_srv  = os.getenv("EMAIL_SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("EMAIL_SMTP_PORT", 587))
    to        = recipient or os.getenv("EMAIL_RECIPIENT", user)

    if not user or not password:
        logger.error("EMAIL_USER ou EMAIL_PASSWORD não configurados.")
        return False

    try:
        xlsx_bytes = build_excel_reports(matches)
        now_str = now_br().strftime("%Y-%m-%d")
        filename = f"gtscout_partidas_{now_str}.xlsx"

        msg = MIMEMultipart()
        msg["From"]    = user
        msg["To"]      = to
        msg["Subject"] = f"📊 GT Scout — Relatório Semanal ({now_str})"

        body = (
            f"Relatório GT Scout\n\n"
            f"Data: {now_br().strftime('%d/%m/%Y %H:%M')}\n"
            f"Total de partidas: {len(matches)}\n\n"
            "GT Scout Bot 🤖"
        )
        msg.attach(MIMEText(body, "plain"))

        part = MIMEBase("application", "octet-stream")
        part.set_payload(xlsx_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

        with smtplib.SMTP(smtp_srv, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(user, password)
            server.sendmail(user, to, msg.as_string())

        logger.info(f"Email enviado para {to}.")
        return True
    except Exception as e:
        logger.error(f"Erro ao enviar email: {e}")
        return False


# ──────────────────────────────────────────────────────────────
# 5. Todos os confrontos (H2H completo — sem filtro de player)
# ──────────────────────────────────────────────────────────────
def build_excel_all_h2h(matches_list) -> bytes:
    """
    Planilha com todos os confrontos de todos os players.
    Cada linha é uma partida com os dois players, placar e vencedor.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Todos os Confrontos"

    headers = [
        "#", "Data", "Temporada",
        "Player Casa", "Gols Casa",
        "Gols Visit.", "Player Visitante",
        "Vencedor"
    ]
    widths = [5, 18, 28, 18, 11, 11, 18, 22]
    _header_row(ws, headers, widths)

    for ri, m in enumerate(matches_list, 2):
        hs, as_ = m.home_score, m.away_score
        fill = _frow(ri)

        if hs is None or as_ is None:
            result_str = "—"
            fh = fa = fill; fnh = fna = FONT_MUTED
        elif hs > as_:
            result_str = m.home_nickname
            fh, fnh = FILL_WIN, FONT_WIN
            fa, fna = FILL_LOSE, FONT_LOSE
        elif hs < as_:
            result_str = m.away_nickname
            fh, fnh = FILL_LOSE, FONT_LOSE
            fa, fna = FILL_WIN, FONT_WIN
        else:
            result_str = "Empate"
            fh = fa = FILL_DRAW; fnh = fna = FONT_DRAW

        kickoff = (m.kickoff or "")[:16].replace("T", " ")
        row = [
            ri - 1, kickoff, m.season_name or "",
            m.home_nickname or "", hs if hs is not None else "",
            as_ if as_ is not None else "", m.away_nickname or "",
            result_str
        ]

        for ci, val in enumerate(row, 1):
            if ci == 4:
                _cell(ws, ri, ci, val, fh, fnh)
            elif ci == 7:
                _cell(ws, ri, ci, val, fa, fna)
            elif ci == 8:
                if result_str == "Empate":
                    _cell(ws, ri, ci, val, FILL_DRAW, FONT_DRAW)
                elif result_str not in ("—", ""):
                    _cell(ws, ri, ci, val, FILL_WIN, FONT_WIN)
                else:
                    _cell(ws, ri, ci, val, fill, FONT_MUTED)
            else:
                _cell(ws, ri, ci, val, fill, FONT_MUTED if ci in (1, 2) else FONT_NORMAL)
        ws.row_dimensions[ri].height = 20

    return _to_bytes(wb)
